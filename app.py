import base64
from datetime import datetime
import io
import os
import re
import qrcode
import streamlit as st
import pytz
from fpdf import FPDF
from openpyxl import Workbook, load_workbook

# Configuración inicial
st.set_page_config(
    page_title="Documento de Identificación (DI) - Residuos",
    page_icon="🚛",
    layout="wide",
    initial_sidebar_state="collapsed"
)

EXCEL_PATH = "registro_documentos.xlsx"
SPAIN_TZ = pytz.timezone("Europe/Madrid")
URL_BASE_APP = os.getenv("URL_BASE_APP", "https://gestor-residuos-di-zv7k5cappd8wle3kzxlxskd.streamlit.app")

# --- BASE DE DATOS DE PROVINCIAS Y MUNICIPIOS/CP ---
LISTA_PROVINCIAS = [
    "", "Álava", "Albacete", "Alicante", "Almería", "Asturias", "Ávila", "Badajoz", "Barcelona",
    "Burgos", "Cáceres", "Cádiz", "Cantabria", "Castellón", "Ciudad Real", "Córdoba", "A Coruña",
    "Cuenca", "Girona", "Granada", "Guadalajara", "Gipuzkoa", "Huelva", "Huesca", "Illes Balears",
    "Jaén", "La Rioja", "Las Palmas", "León", "Lleida", "Lugo", "Madrid", "Málaga", "Murcia",
    "Navarra", "Ourense", "Palencia", "Pontevedra", "Salamanca", "Santa Cruz de Tenerife", "Segovia",
    "Sevilla", "Soria", "Tarragona", "Teruel", "Toledo", "Valencia", "Valladolid", "Bizkaia",
    "Zamora", "Zaragoza", "Ceuta", "Melilla"
]

# Prefijos de Código Postal por Provincia
PREFIJOS_CP = {
    "Álava": "01", "Albacete": "02", "Alicante": "03", "Almería": "04", "Ávila": "05",
    "Badajoz": "06", "Illes Balears": "07", "Barcelona": "08", "Burgos": "09", "Cáceres": "10",
    "Cádiz": "11", "Castellón": "12", "Ciudad Real": "13", "Córdoba": "14", "A Coruña": "15",
    "Cuenca": "16", "Girona": "17", "Granada": "18", "Guadalajara": "19", "Gipuzkoa": "20",
    "Huelva": "21", "Huesca": "22", "Jaén": "23", "León": "24", "Lleida": "25", "La Rioja": "26",
    "Lugo": "27", "Madrid": "28", "Málaga": "29", "Murcia": "30", "Navarra": "31", "Ourense": "32",
    "Asturias": "33", "Palencia": "34", "Las Palmas": "35", "Pontevedra": "36", "Salamanca": "37",
    "Santa Cruz de Tenerife": "38", "Cantabria": "39", "Segovia": "40", "Sevilla": "41", "Soria": "42",
    "Tarragona": "43", "Teruel": "44", "Toledo": "45", "Valencia": "46", "Valladolid": "47",
    "Bizkaia": "48", "Zamora": "49", "Zaragoza": "50", "Ceuta": "51", "Melilla": "52"
}

# Municipios de referencia (puedes ampliar esta lista)
MUNICIPIOS_POR_PROVINCIA = {
    "Granada": ["Granada", "Motril", "Almuñécar", "Armilla", "Baza", "Iznalloz", "Loja", "Maracena"],
    "Málaga": ["Málaga", "Marbella", "Mijas", "Fuengirola", "Vélez-Málaga", "Estepona", "Torremolinos", "Antequera"],
    "Sevilla": ["Sevilla", "Dos Hermanas", "Alcalá de Guadaíra", "Utrera", "Ecija", "Mairena del Aljarafe"],
    "Madrid": ["Madrid", "Móstoles", "Alcalá de Henares", "Fuenlabrada", "Leganés", "Getafe", "Alcorcón"],
    "Barcelona": ["Barcelona", "L'Hospitalet de Llobregat", "Badalona", "Terrassa", "Sabadell", "Mataró"],
    "Valencia": ["Valencia", "Torrent", "Gandia", "Paterna", "Sagunto"],
}

# --- FUNCIONES DE VALIDACIÓN ---
def validar_nif_cif_nie(documento: str) -> bool:
    doc = documento.strip().upper()
    if not doc:
        return True
    pattern = r'^([ABCDEFGHJKLMNPQRSUVW]\d{7}[0-9A-J]|[XYZ]\d{7}[A-Z]|\d{8}[A-Z])$'
    return bool(re.match(pattern, doc))

def validar_cp(cp: str) -> bool:
    cp_clean = cp.strip()
    if not cp_clean:
        return True
    return len(cp_clean) == 5 and cp_clean.isdigit()

def validar_nima(nima: str) -> bool:
    nima_clean = nima.strip()
    if not nima_clean:
        return True
    return len(nima_clean) == 10 and nima_clean.isdigit()

def obtener_ahora_espana():
    return datetime.now(SPAIN_TZ)

def obtener_siguiente_correlativo() -> int:
    if not os.path.exists(EXCEL_PATH):
        return 1
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        return max(1, ws.max_row)
    except Exception:
        return 1

def generar_numero_di(nima_operador: str, correlativo: int) -> str:
    ahora = obtener_ahora_espana()
    nima_limpio = nima_operador.strip() if nima_operador else "0"
    nima_10 = nima_limpio[:10] if len(nima_limpio) >= 10 else nima_limpio.zfill(10)
    anio = ahora.strftime("%Y")
    correlativo_str = str(correlativo).zfill(3)
    return f"{nima_10}{anio}{correlativo_str}"

# OPCIONES
OPCIONES_OPERADOR = ["A02", "P03 (Productor > 1000 Tn RNP)", "P04 (Productor < 1000 Tn RNP)", "G04 (Gestor RNP)", "G05 (Gestor Intermedio RNP)"]
OPCIONES_ORIGEN = ["P03 (Productor > 1000 Tn RNP)", "P04 (Productor < 1000 Tn RNP)", "G04 (Gestor RNP)", "G05 (Gestor Intermedio RNP)"]
OPCIONES_DESTINO = ["G04 (Gestor RNP)", "G05 (Gestor Intermedio RNP)"]
OPCIONES_TRANSPORTISTA = ["T02 (Transportista RNP)", "T01", "T03"]

# --- BARRA LATERAL ---
with st.sidebar:
    st.header("📊 Gestión de Registros")
    if os.path.exists(EXCEL_PATH):
        with open(EXCEL_PATH, "rb") as f_excel:
            st.download_button(
                label="📥 Descargar Registro Excel Completo",
                data=f_excel,
                file_name="registro_documentos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="btn_excel_sidebar"
            )
    else:
        st.info("ℹ️ Aún no hay registros guardados.")

# --- MODO VISOR QR ---
if "doc" in st.query_params:
    doc_id = st.query_params["doc"]
    st.title(f"🔎 Verificación de Documento: {doc_id}")
    pdf_filename = f"DI_{doc_id.replace('/', '_')}.pdf"

    if os.path.exists(pdf_filename):
        st.success("✅ Documento original encontrado y verificado.")
        with open(pdf_filename, "rb") as f:
            pdf_bytes = f.read()

        st.download_button("📥 Descargar Documento PDF Oficial", data=pdf_bytes, file_name=pdf_filename, mime="application/pdf", key="btn_visor_pdf")
        st.markdown("---")
        base64_pdf = base64.b64encode(pdf_bytes).decode("utf-8")
        st.markdown(f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="600" type="application/pdf"></iframe>', unsafe_allow_html=True)
    else:
        st.error("❌ El documento no se encuentra en el servidor.")

    if st.button("⬅️ Volver a la aplicación principal"):
        st.query_params.clear()
        st.rerun()
    st.stop()

# --- FORMULARIO PRINCIPAL ---
ahora_espana = obtener_ahora_espana()
st.title("🚛 Documento de Identificación de Residuos (DI) y Carta de Porte")
st.write("Rellena las secciones para generar el PDF oficial y volcar el registro en Excel.")

# 1. IDENTIFICACIÓN DEL DOCUMENTO Y FECHA
st.header("1. IDENTIFICACIÓN DEL DOCUMENTO Y FECHA")
siguiente_correlativo = obtener_siguiente_correlativo()
di_sugerido = generar_numero_di("123456789", siguiente_correlativo)

col_d1, col_d2, col_d3 = st.columns(3)
with col_d1:
    di_num = st.text_input("🆔 Nº Documento (Autogenerado):", value=di_sugerido)
with col_d2:
    fecha_inicio = st.text_input("Fecha inicio traslado:", value=ahora_espana.strftime("%d/%m/%Y"))
with col_d3:
    hora_inicio = st.text_input("Hora (España):", value=ahora_espana.strftime("%H:%M"))

st.markdown("---")

# HELPER PARA COMPONENTE DE DIRECCIÓN INTERACTIVO (Limpio, sin cuadro personalizado)
def selector_ubicacion(prefix_key: str, label_titulo: str):
    col1, col2, col3 = st.columns(3)
    
    with col1:
        prov_sel = st.selectbox(
            f"Provincia {label_titulo}:",
            options=LISTA_PROVINCIAS,
            index=0,
            key=f"{prefix_key}_prov"
        )
    
    muni_options = [""] + MUNICIPIOS_POR_PROVINCIA.get(prov_sel, []) if prov_sel else [""]
    
    with col2:
        muni_sel = st.selectbox(
            f"Municipio {label_titulo}:",
            options=muni_options,
            index=0,
            key=f"{prefix_key}_muni"
        )

    prefijo_cp = PREFIJOS_CP.get(prov_sel, "")
    cp_sugerido = f"{prefijo_cp}001" if prefijo_cp else ""

    with col3:
        cp_sel = st.text_input(
            f"C.P. {label_titulo} (5 dígitos):",
            value=cp_sugerido,
            max_chars=5,
            key=f"{prefix_key}_cp",
            placeholder="Ej: 18001"
        )
    return prov_sel, muni_sel, cp_sel

# 2. OPERADOR DEL TRASLADO
st.header("2. OPERADOR DEL TRASLADO")
c_op1, c_op2 = st.columns(2)
with c_op1:
    op_nif = st.text_input("NIF/CIF Operador:", value="", placeholder="Ej: B12345678")
    op_nombre = st.text_input("Razón Social / Nombre Operador:", value="")
    op_tipo = st.selectbox("Tipo Operador:", OPCIONES_OPERADOR)

req_op = " *" if "P04" not in op_tipo else ""
with c_op2:
    op_nima = st.text_input(f"NIMA Operador (10 dígitos){req_op}:", value="", max_chars=10, placeholder="Ej: 0123456789")
    op_inscripcion = st.text_input(f"Nº Inscripción{req_op}:", value="")
    op_direccion = st.text_input("Dirección Operador:", value="")

op_prov, op_muni, op_cp = selector_ubicacion("op", "Operador")

col_op_extra1, col_op_extra2 = st.columns(2)
with col_op_extra1:
    op_telefono = st.text_input("Teléfono Operador:", value="")
with col_op_extra2:
    op_email = st.text_input("Correo Electrónico Operador:", value="")

st.markdown("---")

# 3. ORIGEN DEL TRASLADO
st.header("3. ORIGEN DEL TRASLADO")
c1, c2 = st.columns(2)
with c1:
    ori_nif = st.text_input("NIF/CIF Origen:", value="", placeholder="Ej: A87654321")
    ori_nombre = st.text_input("Razón Social Origen:", value="")
    ori_tipo = st.selectbox("Tipo Origen:", OPCIONES_ORIGEN)

req_ori = " *" if "P04" not in ori_tipo else ""
with c2:
    ori_nima = st.text_input(f"NIMA Origen (10 dígitos){req_ori}:", value="", max_chars=10, placeholder="Ej: 0123456789")
    ori_inscripcion = st.text_input(f"Nº Inscripción Origen{req_ori}:", value="")
    ori_direccion = st.text_input("Dirección Origen:", value="")

ori_prov, ori_muni, ori_cp = selector_ubicacion("ori", "Origen")

col_ori_extra1, col_ori_extra2 = st.columns(2)
with col_ori_extra1:
    ori_telefono = st.text_input("Teléfono Origen:", value="")
with col_ori_extra2:
    ori_email = st.text_input("Email Origen:", value="")

st.markdown("---")

# 4. DESTINO DEL TRASLADO
st.header("4. DESTINO DEL TRASLADO")
c1, c2 = st.columns(2)
with c1:
    des_nif = st.text_input("NIF/CIF Destino:", value="", placeholder="Ej: B99887766")
    des_nombre = st.text_input("Razón Social Destino:", value="")
    des_tipo = st.selectbox("Tipo Destino:", OPCIONES_DESTINO)

req_des = " *" if "P04" not in des_tipo else ""
with c2:
    des_nima = st.text_input(f"NIMA Destino (10 dígitos){req_des}:", value="", max_chars=10, placeholder="Ej: 0123456789")
    des_inscripcion = st.text_input(f"Nº Inscripción Destino{req_des}:", value="")
    des_direccion = st.text_input("Dirección Destino:", value="")

des_prov, des_muni, des_cp = selector_ubicacion("des", "Destino")

col_des_extra1, col_des_extra2 = st.columns(2)
with col_des_extra1:
    des_telefono = st.text_input("Teléfono Destino:", value="")
with col_des_extra2:
    des_email = st.text_input("Email Destino:", value="")

st.markdown("---")

# 5. INFORMACIÓN SOBRE EL RESIDUO
st.header("5. INFORMACIÓN SOBRE EL RESIDUO QUE SE TRASLADA")
c1, c2 = st.columns(2)
with c1:
    ler = st.text_input("Código LER:", value="")
    desc_residuo = st.text_area("Descripción del residuo:", value="")
    cantidad_kg = st.text_input("Cantidad (kg):", value="")
with c2:
    operacion_tratam = st.text_input("Operación Tratamiento Destino:", value="")
    operacion_desagregada = st.text_input("Operación Destino Desagregada:", value="")
    desc_operacion = st.text_input("Descripción Op. Tratamiento:", value="")

st.markdown("---")

# 6. INFORMACIÓN RELATIVA AL TRANSPORTISTA (Con Dirección completa, Provincia, Municipio y CP)
st.header("6. INFORMACIÓN RELATIVA AL TRANSPORTISTA")
c1, c2 = st.columns(2)
with c1:
    trans_nif = st.text_input("N.I.F./CIF Transportista:", value="")
    trans_nombre = st.text_input("Razón Social / Nombre Transportista:", value="")
    trans_tipo = st.selectbox("Tipo Transportista:", OPCIONES_TRANSPORTISTA, index=0)

req_trans = " *" if "P04" not in trans_tipo else ""
with c2:
    trans_nima = st.text_input(f"NIMA Transportista (10 dígitos){req_trans}:", value="", max_chars=10, placeholder="Ej: 0123456789")
    trans_inscripcion = st.text_input(f"Nº Inscripción / Autorización{req_trans}:", value="")
    trans_direccion = st.text_input("Dirección Transportista:", value="")

trans_prov, trans_muni, trans_cp = selector_ubicacion("trans", "Transportista")

c3, c4 = st.columns(2)
with c3:
    trans_conductor = st.text_input("Conductor:", value="")
    trans_matricula = st.text_input("Matrícula y Vehículo:", value="")
with c4:
    trans_telefono = st.text_input("Teléfono Transportista:", value="")
    trans_email = st.text_input("Email Transportista:", value="")

st.markdown("---")

# 7. ACEPTACIÓN DEL RESIDUO
st.header("7. INFORMACIÓN SOBRE LA ACEPTACIÓN DEL RESIDUO")
c1, c2, c3 = st.columns(3)
with c1:
    fecha_entrega = st.text_input("Fecha Entrega:", value="")
    kg_recibidos = st.text_input("Kg. Netos Recibidos:", value="")
with c2:
    fecha_aceptacion = st.text_input("Fecha Aceptación/Rechazo:", value="")
    aceptacion_estado = st.selectbox("Aceptación:", ["", "Sí", "No"], index=0)
with c3:
    motivo_rechazo = st.text_input("Motivo de rechazo (si aplica):", value="")

st.markdown("---")
btn_generar = st.button("🚀 Generar PDF Oficial y Registrar", type="primary")

# --- PROCESAMIENTO Y VALIDACIONES ---
if btn_generar:
    errores = []

    if not di_num:
        errores.append("El Número de Documento (DI) es obligatorio.")

    # Validar Formatos de NIF / CIF / NIE
    for campo, val in [("Operador", op_nif), ("Origen", ori_nif), ("Destino", des_nif), ("Transportista", trans_nif)]:
        if val and not validar_nif_cif_nie(val):
            errores.append(f"El NIF/CIF/NIE del {campo} ('{val}') no tiene un formato válido.")

    # Validar Formatos de Código Postal
    for campo, val in [("Operador", op_cp), ("Origen", ori_cp), ("Destino", des_cp), ("Transportista", trans_cp)]:
        if val and not validar_cp(val):
            errores.append(f"El Código Postal del {campo} debe ser un número de 5 dígitos.")

    # Validar Formatos de NIMA
    for campo, val in [("Operador", op_nima), ("Origen", ori_nima), ("Destino", des_nima), ("Transportista", trans_nima)]:
        if val and not validar_nima(val):
            errores.append(f"El NIMA del {campo} ('{val}') debe tener exactamente 10 dígitos numéricos.")

    # Validaciones condicionales (si NO es P04)
    if "P04" not in op_tipo:
        if not op_nima.strip():
            errores.append("El NIMA del Operador es obligatorio para el tipo seleccionado.")
        if not op_inscripcion.strip():
            errores.append("El Nº de Inscripción del Operador es obligatorio para el tipo seleccionado.")

    if "P04" not in ori_tipo:
        if not ori_nima.strip():
            errores.append("El NIMA de Origen es obligatorio para el tipo seleccionado.")
        if not ori_inscripcion.strip():
            errores.append("El Nº de Inscripción de Origen es obligatorio para el tipo seleccionado.")

    if "P04" not in des_tipo:
        if not des_nima.strip():
            errores.append("El NIMA de Destino es obligatorio para el tipo seleccionado.")
        if not des_inscripcion.strip():
            errores.append("El Nº de Inscripción de Destino es obligatorio para el tipo seleccionado.")

    if "P04" not in trans_tipo:
        if not trans_nima.strip():
            errores.append("El NIMA del Transportista es obligatorio para el tipo seleccionado.")
        if not trans_inscripcion.strip():
            errores.append("El Nº de Inscripción del Transportista es obligatorio para el tipo seleccionado.")

    if errores:
        for err in errores:
            st.error(f"⚠️ {err}")
    else:
        # Generación del QR, PDF y Excel
        base_limpia = URL_BASE_APP.strip().rstrip("/")
        enlace_qr = f"{base_limpia}?doc={di_num}"

        qr = qrcode.QRCode(box_size=10, border=2)
        qr.add_data(enlace_qr)
        qr.make(fit=True)
        img_qr = qr.make_image(fill_color="black", back_color="white")
        qr_path = "temp_qr.png"
        img_qr.save(qr_path)

        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=10)

        def s(txt):
            return str(txt).encode("latin-1", "replace").decode("latin-1")

        pdf.image(qr_path, x=150, y=10, w=50, h=50)

        pdf.set_font("Arial", "B", 10)
        pdf.cell(135, 6, s("DOCUMENTO DE IDENTIFICACIÓN DE RESIDUOS Y CARTA DE PORTE"), border=0, ln=True)
        pdf.ln(2)
        pdf.set_font("Arial", "", 8)
        pdf.cell(135, 6, s(f"Documento de Identificación nº: {di_num}"), border=1, ln=True)
        pdf.cell(67, 6, s(f"Fecha inicio traslado: {fecha_inicio}"), border=1)
        pdf.cell(68, 6, s(f"Hora: {hora_inicio}"), border=1, ln=True)

        pdf.set_y(63)

        # Operador
        pdf.set_font("Arial", "B", 8)
        pdf.set_fill_color(220, 220, 220)
        pdf.cell(190, 5, s("OPERADOR DEL TRASLADO"), border=1, ln=True, fill=True)
        pdf.set_font("Arial", "", 7.5)
        pdf.cell(50, 5, s(f"NIF: {op_nif}"), border=1)
        pdf.cell(140, 5, s(f"Razón social/Nombre: {op_nombre}"), border=1, ln=True)
        pdf.cell(50, 5, s(f"NIMA: {op_nima}"), border=1)
        pdf.cell(80, 5, s(f"Nº inscripción: {op_inscripcion}"), border=1)
        pdf.cell(60, 5, s(f"Tipo: {op_tipo}"), border=1, ln=True)
        pdf.cell(130, 5, s(f"Dirección: {op_direccion}"), border=1)
        pdf.cell(60, 5, s(f"C.P.: {op_cp}"), border=1, ln=True)
        pdf.cell(95, 5, s(f"Municipio: {op_muni}"), border=1)
        pdf.cell(95, 5, s(f"Provincia: {op_prov}"), border=1, ln=True)
        pdf.cell(95, 5, s(f"Teléfono: {op_telefono}"), border=1)
        pdf.cell(95, 5, s(f"Correo electrónico: {op_email}"), border=1, ln=True)
        pdf.ln(3)

        # Origen
        pdf.set_font("Arial", "B", 8)
        pdf.cell(190, 5, s("ORIGEN DEL TRASLADO"), border=1, ln=True, fill=True)
        pdf.set_font("Arial", "I", 7.5)
        pdf.cell(190, 4, s("Información de la instalación origen del traslado:"), border="LR", ln=True)
        pdf.set_font("Arial", "", 7.5)
        pdf.cell(50, 5, s(f"NIF: {ori_nif}"), border=1)
        pdf.cell(140, 5, s(f"Razón social/Nombre: {ori_nombre}"), border=1, ln=True)
        pdf.cell(50, 5, s(f"NIMA: {ori_nima}"), border=1)
        pdf.cell(80, 5, s(f"Nº inscripción: {ori_inscripcion}"), border=1)
        pdf.cell(60, 5, s(f"Tipo: {ori_tipo}"), border=1, ln=True)
        pdf.cell(130, 5, s(f"Dirección: {ori_direccion}"), border=1)
        pdf.cell(60, 5, s(f"C.P.: {ori_cp}"), border=1, ln=True)
        pdf.cell(95, 5, s(f"Municipio: {ori_muni}"), border=1)
        pdf.cell(95, 5, s(f"Provincia: {ori_prov}"), border=1, ln=True)
        pdf.cell(95, 5, s(f"Teléfono: {ori_telefono}"), border=1)
        pdf.cell(95, 5, s(f"Correo electrónico: {ori_email}"), border=1, ln=True)
        pdf.ln(3)

        # Destino
        pdf.set_font("Arial", "B", 8)
        pdf.cell(190, 5, s("DESTINO DEL TRASLADO"), border=1, ln=True, fill=True)
        pdf.set_font("Arial", "I", 7.5)
        pdf.cell(190, 4, s("Información de la instalación de destino:"), border="LR", ln=True)
        pdf.set_font("Arial", "", 7.5)
        pdf.cell(50, 5, s(f"NIF: {des_nif}"), border=1)
        pdf.cell(140, 5, s(f"Razón social/Nombre: {des_nombre}"), border=1, ln=True)
        pdf.cell(50, 5, s(f"NIMA: {des_nima}"), border=1)
        pdf.cell(80, 5, s(f"Nº inscripción: {des_inscripcion}"), border=1)
        pdf.cell(60, 5, s(f"Tipo: {des_tipo}"), border=1, ln=True)
        pdf.cell(130, 5, s(f"Dirección: {des_direccion}"), border=1)
        pdf.cell(60, 5, s(f"C.P.: {des_cp}"), border=1, ln=True)
        pdf.cell(95, 5, s(f"Municipio: {des_muni}"), border=1)
        pdf.cell(95, 5, s(f"Provincia: {des_prov}"), border=1, ln=True)
        pdf.cell(95, 5, s(f"Teléfono: {des_telefono}"), border=1)
        pdf.cell(95, 5, s(f"Correo electrónico: {des_email}"), border=1, ln=True)
        pdf.ln(3)

        # Residuo
        pdf.set_font("Arial", "B", 8)
        pdf.cell(190, 5, s("INFORMACIÓN SOBRE EL RESIDUO QUE SE TRASLADA"), border=1, ln=True, fill=True)
        pdf.set_font("Arial", "", 7.5)
        pdf.cell(50, 5, s(f"Código LER: {ler}"), border=1)
        pdf.cell(140, 5, s(f"Descripción: {desc_residuo}"), border=1, ln=True)
        pdf.cell(95, 5, s(f"Op. Tratamiento Destino: {operacion_tratam}"), border=1)
        pdf.cell(95, 5, s(f"Op. Tratamiento Desagregada: {operacion_desagregada}"), border=1, ln=True)
        pdf.cell(130, 5, s(f"Descripción Op. Tratamiento: {desc_operacion}"), border=1)
        pdf.cell(60, 5, s(f"Cantidad (kg): {cantidad_kg}"), border=1, ln=True)
        pdf.ln(3)

        # Transportista
        pdf.set_font("Arial", "B", 8)
        pdf.cell(190, 5, s("INFORMACIÓN RELATIVA AL TRANSPORTISTA"), border=1, ln=True, fill=True)
        pdf.set_font("Arial", "", 7.5)
        pdf.cell(50, 5, s(f"N.I.F.: {trans_nif}"), border=1)
        pdf.cell(140, 5, s(f"Razón social/Nombre: {trans_nombre}"), border=1, ln=True)
        pdf.cell(50, 5, s(f"NIMA: {trans_nima}"), border=1)
        pdf.cell(80, 5, s(f"Nº inscripción: {trans_inscripcion}"), border=1)
        pdf.cell(60, 5, s(f"Tipo: {trans_tipo}"), border=1, ln=True)
        pdf.cell(130, 5, s(f"Dirección: {trans_direccion}"), border=1)
        pdf.cell(60, 5, s(f"C.P.: {trans_cp}"), border=1, ln=True)
        pdf.cell(95, 5, s(f"Municipio: {trans_muni}"), border=1)
        pdf.cell(95, 5, s(f"Provincia: {trans_prov}"), border=1, ln=True)
        pdf.cell(95, 5, s(f"Conductor: {trans_conductor}"), border=1)
        pdf.cell(95, 5, s(f"Matrícula: {trans_matricula}"), border=1, ln=True)
        pdf.cell(95, 5, s(f"Teléfono: {trans_telefono}"), border=1)
        pdf.cell(95, 5, s(f"Email: {trans_email}"), border=1, ln=True)
        pdf.ln(3)

        # Aceptación
        pdf.set_font("Arial", "B", 8)
        pdf.cell(190, 5, s("INFORMACIÓN SOBRE LA ACEPTACIÓN DEL RESIDUO"), border=1, ln=True, fill=True)
        pdf.set_font("Arial", "", 7.5)
        pdf.cell(95, 5, s(f"Fecha entrega: {fecha_entrega}"), border=1)
        pdf.cell(95, 5, s(f"Kg. netos recibidos: {kg_recibidos}"), border=1, ln=True)
        pdf.cell(95, 5, s(f"Fecha aceptación/rechazo: {fecha_aceptacion}"), border=1)
        pdf.cell(95, 5, s(f"Aceptación: [{aceptacion_estado}]"), border=1, ln=True)
        if motivo_rechazo:
            pdf.cell(190, 5, s(f"Motivo de rechazo: {motivo_rechazo}"), border=1, ln=True)

        pdf_out_filename = f"DI_{di_num.replace('/', '_')}.pdf"
        pdf.output(pdf_out_filename)
        pdf_bytes = pdf.output(dest='S').encode('latin-1')

        # Registro en Excel
        columnas_excel = [
            "Nº DI", "Fecha Inicio Traslado", "Hora Inicio", "NIF Operador", "Razón Social Operador",
            "NIMA Operador", "Nº Inscripción Operador", "Tipo Operador", "Dirección Operador", "CP Operador",
            "Municipio Operador", "Provincia Operador", "Teléfono Operador", "Email Operador",
            "NIF Origen", "Razón Social Origen", "NIMA Origen", "Nº Inscripción Origen", "Tipo Origen",
            "Dirección Origen", "CP Origen", "Municipio Origen", "Provincia Origen", "Teléfono Origen",
            "Email Origen", "NIF Destino", "Razón Social Destino", "NIMA Destino", "Nº Inscripción Destino",
            "Tipo Destino", "Dirección Destino", "CP Destino", "Municipio Destino", "Provincia Destino",
            "Teléfono Destino", "Email Destino", "Código LER", "Descripción Residuo", "Cantidad (kg)",
            "Op. Tratamiento Destino", "Op. Tratamiento Desagregada", "Descripción Op. Tratamiento",
            "NIF Transportista", "Razón Social Transportista", "NIMA Transportista", "Nº Inscripción Transportista",
            "Tipo Transportista", "Dirección Transportista", "CP Transportista", "Municipio Transportista", 
            "Provincia Transportista", "Conductor", "Matrícula / Vehículo", "Teléfono Transportista", 
            "Email Transportista", "Fecha Entrega", "Kg Netos Recibidos", "Fecha Aceptación/Rechazo", 
            "Estado Aceptación", "Motivo Rechazo", "Enlace QR Verificación"
        ]

        fila_datos = [
            di_num, fecha_inicio, hora_inicio, op_nif, op_nombre,
            op_nima, op_inscripcion, op_tipo, op_direccion, op_cp,
            op_muni, op_prov, op_telefono, op_email,
            ori_nif, ori_nombre, ori_nima, ori_inscripcion, ori_tipo,
            ori_direccion, ori_cp, ori_muni, ori_prov, ori_telefono,
            ori_email, des_nif, des_nombre, des_nima, des_inscripcion,
            des_tipo, des_direccion, des_cp, des_muni, des_prov,
            des_telefono, des_email, ler, desc_residuo, cantidad_kg,
            operacion_tratam, operacion_desagregada, desc_operacion,
            trans_nif, trans_nombre, trans_nima, trans_inscripcion,
            trans_tipo, trans_direccion, trans_cp, trans_muni,
            trans_prov, trans_conductor, trans_matricula, trans_telefono, 
            trans_email, fecha_entrega, kg_recibidos, fecha_aceptacion, 
            aceptacion_estado, motivo_rechazo, enlace_qr
        ]

        if not os.path.exists(EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = "Registros DI"
            ws.append(columnas_excel)
        else:
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active

        ws.append(fila_datos)
        wb.save(EXCEL_PATH)

        st.success(f"✅ Documento generado y registrado correctamente: **{di_num}**")

        col_a, col_b = st.columns([1, 3])
        with col_a:
            st.image(qr_path, caption="Código QR Generado", width=150)
        with col_b:
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                st.download_button("📄 Descargar PDF Oficial", data=pdf_bytes, file_name=pdf_out_filename, mime="application/pdf", key="btn_pdf_main")
            with col_btn2:
                with open(EXCEL_PATH, "rb") as f_excel:
                    st.download_button("📊 Descargar Registro Excel", data=f_excel, file_name="registro_documentos.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="btn_excel_main")
